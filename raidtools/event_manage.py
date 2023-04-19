import logging
from io import BytesIO
from typing import Dict

import discord.ui
from openpyxl.styles import PatternFill
from openpyxl.workbook import Workbook

from raidtools.confirmation import DeleteConfirmationView
from raidtools.discordevent import RaidtoolsDiscordEvent
from raidtools.shared import EventEmbed

log = logging.getLogger("red.karlo-cogs.raidtools")


class EventManageView(discord.ui.View):
    def __init__(self, config, events):
        self.config = config
        self.events = events
        super().__init__()
        self.add_item(EventManageDropdown(self.config, events=events))


class EventManageDropdown(discord.ui.Select):
    def __init__(self, config, events):
        self.config = config
        self.events = events
        options = [
            discord.SelectOption(
                label=event["event_name"], description=f"ID: {event_id}", value=event_id
            )
            for event_id, event in events.items()
        ]

        super().__init__(placeholder="Odaberi event", options=options)

    async def callback(self, interaction: discord.Interaction) -> None:
        event_id = self.values[0]
        event = self.events[event_id]
        embed = await EventEmbed.create_event_embed(
            signed_up=event["signed_up"],
            event_info=event,
            bot=interaction.client,
            config=self.config,
        )
        await interaction.response.send_message(
            content="Uređuješ ovaj event:",
            embed=embed,
            view=EventEditView(
                event_id=event_id, config=self.config, preview_msg=interaction.message
            ),
            ephemeral=True,
        )


class EventEditView(discord.ui.View):
    def __init__(self, event_id, config, preview_msg: discord.Message):
        self.event_id = event_id
        self.config = config
        super().__init__()
        self.add_item(EventEditDropdown(event_id=event_id, config=config, preview_msg=preview_msg))

    @discord.ui.button(label="Izbriši event", style=discord.ButtonStyle.danger, row=1)
    async def delete_event(
        self, interaction: discord.Interaction, button: discord.ui.Button, /
    ) -> None:
        await interaction.response.send_message(
            "Jesi siguran?",
            ephemeral=True,
            view=DeleteConfirmationView(self.config, self.event_id),
        )

    @discord.ui.button(label="Export eventa", style=discord.ButtonStyle.success, row=1)
    async def export_event(
        self, interaction: discord.Interaction, button: discord.ui.Button, /
    ) -> None:
        events: Dict = await self.config.guild(interaction.guild).events()
        event = events[self.event_id]

        all_signups: Dict[str, str] = {}
        for player_class, signups in event["signed_up"].items():
            player_class = player_class.replace("_", " ").title()
            if player_class not in [
                "Warrior",
                "Rogue",
                "Mage",
                "Priest",
                "Druid",
                "Hunter",
                "Shaman",
                "Warlock",
                "Paladin",
                "Monk",
                "Death Knight",
                "Demon Hunter",
                "Evoker",
            ]:
                continue
            for signup in signups:
                all_signups[str(signup)] = player_class

        wb = Workbook()
        ws = wb.active

        ws["A4"] = "Grupa 1"
        ws["B4"] = "Grupa 2"
        ws["E4"] = "Prijave"
        for i, (signup_id, player_class) in enumerate(all_signups.items(), start=5):
            display_name = interaction.guild.get_member(int(signup_id)).display_name
            color = self.get_class_color(player_class)
            ws.cell(row=i, column=5, value=display_name).fill = PatternFill(
                start_color=color, fill_type="solid"
            )

        wb_bytesio = BytesIO()
        wb.save(wb_bytesio)
        wb_bytesio.seek(0)

        file = discord.File(wb_bytesio, filename=f"{event['event_name']}.xlsx")
        await interaction.response.send_message(
            ephemeral=True,
            file=file,
        )

    @staticmethod
    def get_class_color(player_class: str) -> str:
        if player_class == "Warrior":
            return "C69B6D"
        elif player_class == "Rogue":
            return "FFF468"
        elif player_class == "Mage":
            return "69CCF0"
        elif player_class == "Priest":
            return "FFFFFF"
        elif player_class == "Druid":
            return "FF7D0A"
        elif player_class == "Hunter":
            return "ABD473"
        elif player_class == "Shaman":
            return "0070DD"
        elif player_class == "Warlock":
            return "8788EE"
        elif player_class == "Paladin":
            return "F48CBA"
        elif player_class == "Monk":
            return "00FF98"
        elif player_class == "Death Knight":
            return "C41E3A"
        elif player_class == "Demon Hunter":
            return "A330C9"
        elif player_class == "Evoker":
            return "33937F"
        else:
            return "FFFFFF"


class EventEditDropdown(discord.ui.Select):
    def __init__(self, event_id, config, preview_msg: discord.Message):
        self.event_id = event_id
        self.config = config
        self.preview_msg = preview_msg
        options = [
            discord.SelectOption(label="Uredi naziv i opis", value="edit_name"),
            discord.SelectOption(label="Uredi vrijeme eventa", value="edit_time"),
        ]

        super().__init__(placeholder="Odaberi radnju", options=options)

    async def callback(self, interaction: discord.Interaction) -> None:
        if self.values[0] == "edit_name":
            await interaction.response.send_modal(
                EventEditNameModal(
                    event_id=self.event_id, config=self.config, preview_msg=self.preview_msg
                )
            )
        elif self.values[0] == "edit_time":
            await interaction.response.send_message(
                "**PROČITAJ ME**\n"
                "Za uređivanje vremena eventa, upiši vremensku oznaku koju možeš dobiti putem "
                "https://r.3v.fi/discord-timestamps/\n"
                "Klikni gumb nakon što imaš vremensku oznaku kopiranu.",
                ephemeral=True,
                view=EventEditTimeView(self.event_id, self.config),
            )


class EventEditNameModal(discord.ui.Modal):
    def __init__(self, event_id, config, preview_msg: discord.Message):
        self.event_id = event_id
        self.config = config
        self.preview_msg = preview_msg
        self.title = "Uredi naziv i opis eventa"
        super().__init__(timeout=600, title=self.title)

        self.event_name = discord.ui.TextInput(
            label="Naziv eventa",
            style=discord.TextStyle.short,
            placeholder="Sepulcher of the First Ones",
            max_length=100,
            required=False,
        )
        self.event_description = discord.ui.TextInput(
            label="Opis eventa",
            style=discord.TextStyle.long,
            placeholder="SotFO HC, ako je moguće, 11/11",
            max_length=300,
            required=False,
        )

        self.add_item(self.event_name)
        self.add_item(self.event_description)

    async def on_submit(self, interaction: discord.Interaction, /) -> None:
        event_name = str(self.event_name)
        event_description = str(self.event_description)

        events: Dict = await self.config.guild(interaction.guild).events()

        event = events[self.event_id]
        if event_name:
            event["event_name"] = event_name
        if event_description:
            event["event_description"] = event_description
        events[self.event_id] = event

        await self.config.guild(interaction.guild).events.set(events)

        event_msg = await self.update_event(event, interaction)

        # Update Discord scheduled event
        scheduled_event = RaidtoolsDiscordEvent(event_msg, interaction, event)
        try:
            await scheduled_event.edit_event()
        except ValueError as e:
            log.error(f"Error while editing scheduled event: {e}")

    async def update_event(self, event, interaction) -> discord.Message:
        embed = await EventEmbed.create_event_embed(
            signed_up=event["signed_up"],
            event_info=event,
            bot=interaction.client,
            config=self.config,
        )

        event_channel = interaction.guild.get_channel_or_thread(event["event_channel"])
        event_msg: discord.Message = await event_channel.fetch_message(event["event_id"])

        await event_msg.edit(embed=embed)
        await interaction.response.send_message(
            "Naziv i opis eventa su promijenjeni.", ephemeral=True
        )
        return event_msg


class EventEditTimeView(discord.ui.View):
    def __init__(self, event_id, config):
        self.event_id = event_id
        self.config = config
        super().__init__()

    @discord.ui.button(label="Klikni me", style=discord.ButtonStyle.primary)
    async def edit_time(
        self, interaction: discord.Interaction, button: discord.ui.Button, /
    ) -> None:
        await interaction.response.send_modal(
            EventEditTimeModal(event_id=self.event_id, config=self.config)
        )


class EventEditTimeModal(discord.ui.Modal):
    def __init__(self, event_id, config):
        self.event_id = event_id
        self.config = config
        self.title = "Uredi vrijeme eventa"
        super().__init__(timeout=600, title=self.title)

        self.event_date = discord.ui.TextInput(
            label="Početak eventa",
            style=discord.TextStyle.short,
            placeholder="<t:1666087800:R> (pročitaj upute)",
            max_length=20,
            required=False,
        )
        self.event_end_date = discord.ui.TextInput(
            label="Kraj eventa",
            style=discord.TextStyle.short,
            placeholder="<t:1823764123:R> (pročitaj upute)",
            max_length=20,
            required=False,
        )

        self.add_item(self.event_date)
        self.add_item(self.event_end_date)

    async def on_submit(self, interaction: discord.Interaction, /) -> None:
        event_date = str(self.event_date)
        event_end_date = str(self.event_end_date)

        events: Dict = await self.config.guild(interaction.guild).events()

        event = events[self.event_id]
        if event_date:
            event["event_date"] = event_date
        if event_end_date:
            event["event_end_date"] = event_end_date
        events[self.event_id] = event

        await self.config.guild(interaction.guild).events.set(events)

        # Update event
        embed = await EventEmbed.create_event_embed(
            signed_up=event["signed_up"],
            event_info=event,
            bot=interaction.client,
            config=self.config,
        )
        event_channel = interaction.guild.get_channel_or_thread(event["event_channel"])
        event_msg = await event_channel.fetch_message(event["event_id"])
        await event_msg.edit(embed=embed)

        await interaction.response.send_message("Vrijeme eventa je promijenjeno.", ephemeral=True)
